# repo for analysis
"""
GitHub .NET Repository Analyzer for AWS-to-GCP Migration
Automates inventory creation by analyzing repo structure, dependencies, and code patterns.
"""

import os
import json
import requests
from datetime import datetime
from typing import List, Dict, Optional
import re
from urllib.parse import urljoin
import base64
import xml.etree.ElementTree as ET

class GitHubRepoAnalyzer:
    def __init__(self, github_token: str, org_or_user: Optional[str] = None):
        """
        Initialize analyzer with GitHub authentication.
        
        Args:
            github_token: GitHub Personal Access Token (PAT)
            org_or_user: GitHub organization or username (optional, can pass per-call)
        """
        self.github_token = github_token
        self.base_url = "https://api.github.com"
        self.headers = {
            "Authorization": f"token {github_token}",
            "Accept": "application/vnd.github.v3+raw"
        }
        self.org_or_user = org_or_user
        self.inventory = []
    
    def get_repo_files(self, repo_owner: str, repo_name: str, path: str = "") -> List[Dict]:
        """
        Fetch file listing for a repository path.
        Returns list of files with metadata.
        """
        url = f"{self.base_url}/repos/{repo_owner}/{repo_name}/contents/{path}"
        try:
            resp = requests.get(url, headers=self.headers, timeout=10)
            resp.raise_for_status()
            return resp.json() if isinstance(resp.json(), list) else [resp.json()]
        except Exception as e:
            print(f"  ⚠ Error fetching files from {repo_name}: {e}")
            return []

    def get_repo_tree(self, repo_owner: str, repo_name: str, branch: str = "main") -> List[Dict]:
        """
        Fetch the entire repository tree recursively in one call.
        This provides file paths without downloading content.
        """
        url = f"{self.base_url}/repos/{repo_owner}/{repo_name}/git/trees/{branch}?recursive=1"
        try:
            resp = requests.get(url, headers=self.headers, timeout=15)
            # If 'main' fails, try 'master'
            if resp.status_code == 404 and branch == "main":
                return self.get_repo_tree(repo_owner, repo_name, branch="master")
            resp.raise_for_status()
            return resp.json().get("tree", [])
        except Exception as e:
            print(f"  ⚠ Error fetching tree for {repo_name}: {e}")
            return []
    
    def get_file_content(self, repo_owner: str, repo_name: str, file_path: str) -> Optional[str]:
        """
        Fetch raw content of a specific file.
        """
        url = f"{self.base_url}/repos/{repo_owner}/{repo_name}/contents/{file_path}"
        try:
            resp = requests.get(url, headers=self.headers, timeout=10)
            if resp.status_code == 200:
                # If it's a JSON response with content, decode base64
                if "content" in resp.json():
                    import base64
                    return base64.b64decode(resp.json()["content"]).decode('utf-8')
                return resp.text
            return None
        except Exception as e:
            print(f"  ⚠ Could not fetch {file_path}: {e}")
            return None
    
    def detect_app_type(self, repo_owner: str, repo_name: str, tree: List[Dict]) -> Dict:
        """
        Detect application type and framework version from tree analysis.
        """
        results = {
            "types": [],
            "framework": "Unknown",
            "suggested_gcp": "Compute Engine (Default)"
        }
        
        # 1. Identify key files from tree
        csproj_files = [f["path"] for f in tree if f["path"].endswith(".csproj")]
        sln_files = [f["path"] for f in tree if f["path"].endswith(".sln")]
        asmx_files = [f["path"] for f in tree if f["path"].endswith(".asmx")]
        web_configs = [f["path"] for f in tree if f["path"].lower().endswith("web.config")]
        
        if asmx_files:
            results["types"].append("SOAP API (ASMX)")
            results["suggested_gcp"] = "Compute Engine / App Engine Flex"

        # 2. Analyze the first .csproj for framework info (Lazy Loading)
        if csproj_files:
            content = self.get_file_content(repo_owner, repo_name, csproj_files[0])
            if content:
                # Detect Framework
                fw_match = re.search(r'<(?:TargetFramework|TargetFrameworkVersion)>(.*?)</', content)
                if fw_match:
                    results["framework"] = fw_match.group(1)
                
                # Detect Type from content
                if "Microsoft.NET.Sdk.Web" in content or "AspNetCore" in content:
                    results["types"].append("REST API (ASP.NET Core)")
                    if "netcore" in results["framework"].lower() or "net5" in results["framework"].lower() or "net6" in results["framework"].lower() or "net7" in results["framework"].lower() or "net8" in results["framework"].lower():
                         results["suggested_gcp"] = "Cloud Run"
                    else:
                         results["suggested_gcp"] = "App Engine Flex / Cloud Run"
                
                if "WindowsService" in content or "TopShelf" in content:
                    results["types"].append("Windows Service")
                    results["suggested_gcp"] = "Compute Engine"
                
                if "Exe" in content and "WindowsService" not in content:
                    results["types"].append("Console App")
                    if results["suggested_gcp"] == "Compute Engine (Default)":
                        results["suggested_gcp"] = "Cloud Run / Batch"

        # 3. Check for web indicators in path
        if any("/Controllers" in f["path"] or "/Views" in f["path"] for f in tree):
            if not any("API" in t for t in results["types"]):
                results["types"].append("Web Application")
        
        return {
            "application_type": "; ".join(results["types"]) if results["types"] else "Library/Other",
            "target_framework": results["framework"],
            "suggested_gcp": results["suggested_gcp"]
        }
    
    def detect_aws_usage(self, repo_owner: str, repo_name: str, tree: List[Dict]) -> Dict:
        """
        Scan for AWS SDK references and configuration patterns in the tree.
        """
        aws_usage = {
            "aws_sdk_detected": False,
            "aws_services": set(),
            "has_hardcoded_aws": False
        }
        
        # High-signal files for AWS usage
        files_to_scan = [f["path"] for f in tree if any(f["path"].endswith(ext) for ext in [".csproj", "packages.config", "web.config", "appsettings.json"])]
        
        for file_path in files_to_scan:
            content = self.get_file_content(repo_owner, repo_name, file_path)
            if not content:
                continue
                
            # Detect SDK packages
            if "AWSSDK" in content or "Amazon." in content:
                aws_usage["aws_sdk_detected"] = True
                services = re.findall(r'(?:AWSSDK|Amazon)\.(\w+)', content)
                aws_usage["aws_services"].update(services)
            
            # Detect hardcoded AWS patterns (Access Keys, Regions)
            if re.search(r'(?:AKIA|ASIA)[0-9A-Z]{16}', content):
                aws_usage["has_hardcoded_aws"] = True
            if "amazonaws.com" in content:
                aws_usage["aws_sdk_detected"] = True

        return {
            "aws_sdk_detected": aws_usage["aws_sdk_detected"],
            "aws_services": ", ".join(sorted(aws_usage["aws_services"])) or ("Yes (Generic)" if aws_usage["aws_sdk_detected"] else "None detected"),
            "has_hardcoded_aws": aws_usage["has_hardcoded_aws"]
        }

    def detect_tech_stack(self, repo_owner: str, repo_name: str, tree: List[Dict]) -> Dict:
        """
        Deep scan for Databases, Auth, and Logging frameworks.
        """
        stack = {
            "database": set(),
            "auth": set(),
            "logging": set(),
            "secrets": "Environment/JSON"
        }
        
        # High-signal files
        files_to_scan = [f["path"] for f in tree if any(f["path"].endswith(ext) for ext in [".csproj", "web.config", "appsettings.json", "app.config"])]
        
        db_patterns = {
            "SQL Server": [r"System\.Data\.SqlClient", r"Microsoft\.Data\.SqlClient", r"SqlServer"],
            "Postgres": [r"Npgsql", r"PostgreSQL"],
            "MySQL": [r"MySql\.Data", r"MySqlConnector"],
            "CosmosDB": [r"Microsoft\.Azure\.Cosmos"],
            "DynamoDB": [r"AWSSDK\.DynamoDBv2"],
            "MongoDB": [r"MongoDB\.Driver"]
        }
        
        auth_patterns = {
            "Windows Auth": [r"WindowsAuthentication", r"Negotiate", r"NTLM"],
            "OAuth/OIDC": [r"OpenIdConnect", r"IdentityServer", r"JwtBearer", r"Microsoft\.Identity"],
            "Identity Framework": [r"Microsoft\.AspNetCore\.Identity", r"AspNet\.Identity"]
        }
        
        logging_patterns = {
            "Serilog": [r"Serilog"],
            "NLog": [r"NLog"],
            "Log4Net": [r"log4net"],
            "Application Insights": [r"ApplicationInsights"]
        }
        
        for file_path in files_to_scan:
            content = self.get_file_content(repo_owner, repo_name, file_path)
            if not content: continue
            
            for db, patterns in db_patterns.items():
                if any(re.search(p, content, re.I) for p in patterns):
                    stack["database"].add(db)
            
            for auth, patterns in auth_patterns.items():
                if any(re.search(p, content, re.I) for p in patterns):
                    stack["auth"].add(auth)
                    
            for log, patterns in logging_patterns.items():
                if any(re.search(p, content, re.I) for p in patterns):
                    stack["logging"].add(log)
            
            if "AzureKeyVault" in content or "AWSSecretsManager" in content:
                stack["secrets"] = "Cloud Secret Manager"

        return {
            "database": ", ".join(sorted(stack["database"])) or "Unknown",
            "auth": ", ".join(sorted(stack["auth"])) or "Default/Basic",
            "logging": ", ".join(sorted(stack["logging"])) or "Standard .NET Logging",
            "secrets_mgmt": stack["secrets"]
        }

    def detect_cicd_pipelines(self, tree: List[Dict]) -> str:
        """
        Identify CI/CD pipelines from file patterns.
        """
        pipelines = []
        paths = [f["path"].lower() for f in tree]
        
        if any(".github/workflows" in p for p in paths):
            pipelines.append("GitHub Actions")
        if any("azure-pipelines" in p for p in paths):
            pipelines.append("Azure Pipelines")
        if any("jenkins" in p or "jenkinsfile" in p for p in paths):
            pipelines.append("Jenkins")
        if any("appveyor.yml" in p for p in paths):
            pipelines.append("AppVeyor")
        if any("bitbucket-pipelines" in p for p in paths):
            pipelines.append("Bitbucket Pipelines")
            
        return ", ".join(pipelines) or "None detected"
    
    def get_repo_metadata(self, repo_owner: str, repo_name: str) -> Optional[Dict]:
        """
        Fetch repository metadata from GitHub API.
        """
        url = f"{self.base_url}/repos/{repo_owner}/{repo_name}"
        try:
            resp = requests.get(url, headers=self.headers, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            return {
                "name": data.get("name"),
                "url": data.get("html_url"),
                "description": data.get("description", ""),
                "stars": data.get("stargazers_count", 0),
                "last_updated": data.get("updated_at"),
                "language": data.get("language"),
                "size_kb": data.get("size", 0),
                "is_fork": data.get("fork", False),
                "is_archived": data.get("archived", False)
            }
        except Exception as e:
            print(f"  ⚠ Error fetching metadata for {repo_name}: {e}")
            return None

    def _parse_repo_spec(self, repo_spec: str) -> Dict:
        """
        Parse repo spec into owner, name, branch, and subpath.
        Supports: "owner/repo" and "owner/repo/tree/branch/subpath"
        """
        parts = repo_spec.strip("/").split("/")
        if len(parts) >= 4 and parts[2] == "tree":
            owner = parts[0]
            name = parts[1]
            branch = parts[3]
            subpath = "/".join(parts[4:])
            return {"owner": owner, "name": name, "branch": branch, "subpath": subpath}
        
        elif len(parts) >= 2:
            return {"owner": parts[0], "name": parts[1], "branch": "main", "subpath": ""}
        
        return {"owner": "", "name": repo_spec, "branch": "main", "subpath": ""}

    def analyze_repo(self, repo_owner: str, repo_name: str, branch: str = "main", subpath: str = "") -> Dict:
        """
        Full analysis of a single repository or subfolder using Tree API.
        """
        display_name = f"{repo_name}/{subpath}" if subpath else repo_name
        print(f"  Analyzing {display_name}...", end="", flush=True)
        
        # 1. Get metadata (Always from root)
        metadata = self.get_repo_metadata(repo_owner, repo_name)
        if not metadata:
            print(" SKIPPED (No metadata)")
            return {}
        
        # 2. Get repository tree (Metadata only)
        tree = self.get_repo_tree(repo_owner, repo_name, branch=branch)
        if not tree:
            print(" SKIPPED (Cannot fetch tree)")
            return {}
        
        # 3. Filter tree if subpath is provided
        if subpath:
            # Ensure subpath ends with / for matching
            match_path = subpath if subpath.endswith("/") else subpath + "/"
            original_size = len(tree)
            tree = [f for f in tree if f["path"].startswith(match_path)]
            if not tree:
                print(f" SKIPPED (Subpath '{subpath}' not found in {branch})")
                return {}
        
        # 4. Detect application type & framework (Lazy downloads)
        app_info = self.detect_app_type(repo_owner, repo_name, tree)
        
        # 5. Detect AWS usage (Lazy downloads)
        aws_usage = self.detect_aws_usage(repo_owner, repo_name, tree)

        # 6. Detect Deep Tech Stack (Lazy downloads)
        tech_stack = self.detect_tech_stack(repo_owner, repo_name, tree)
        
        # 7. Detect CI/CD (Tree only)
        cicd = self.detect_cicd_pipelines(tree)
        
        # 8. Estimate complexity
        # Use a localized size estimate if in a subfolder
        complexity = self._estimate_complexity(app_info["application_type"], aws_usage["aws_sdk_detected"], metadata["size_kb"] if not subpath else 5000)
        
        # 9. Refine Suggested GCP Service
        suggested_gcp = app_info["suggested_gcp"]
        if "Windows Auth" in tech_stack["auth"]:
             suggested_gcp = "Compute Engine (Windows AD required)"
        elif app_info["target_framework"].startswith("netcore") or app_info["target_framework"].startswith("net5") or app_info["target_framework"].startswith("net6") or app_info["target_framework"].startswith("net7") or app_info["target_framework"].startswith("net8"):
             if "Cloud Run" not in suggested_gcp:
                 suggested_gcp = "Cloud Run"

        result = {
            "repo_name": display_name,
            "repo_url": f"{metadata['url']}/tree/{branch}/{subpath}" if subpath else metadata["url"],
            "description": metadata.get("description", ""),
            "application_type": app_info["application_type"],
            "target_framework": app_info["target_framework"],
            "suggested_gcp": suggested_gcp,
            "aws_sdk_detected": aws_usage["aws_sdk_detected"],
            "aws_services": aws_usage["aws_services"],
            "has_hardcoded_aws": aws_usage["has_hardcoded_aws"],
            "database_engine": tech_stack["database"],
            "auth_strategy": tech_stack["auth"],
            "logging_framework": tech_stack["logging"],
            "secrets_mgmt": tech_stack["secrets_mgmt"],
            "cicd_pipeline": cicd,
            "size_kb": metadata["size_kb"],
            "last_updated": metadata.get("last_updated", ""),
            "is_archived": metadata.get("is_archived", False),
            "migration_complexity": complexity,
            "notes": self._generate_notes(app_info["application_type"], aws_usage["aws_sdk_detected"])
        }
        
        if subpath:
            result["notes"] = f"(Subfolder Analysis) {result['notes']}"
        
        print(" ✓")
        return result
    
    def _estimate_complexity(self, app_type: str, has_aws: bool, size_kb: int) -> str:
        """
        Rough complexity estimation for migration.
        """
        score = 0
        
        # Type complexity
        if "Windows Service" in app_type:
            score += 3
        elif "REST API" in app_type or "SOAP" in app_type:
            score += 2
        elif "DLL" in app_type:
            score += 1
        
        # AWS usage
        if has_aws:
            score += 2
        
        # Size
        if size_kb > 10000:
            score += 2
        elif size_kb > 1000:
            score += 1
        
        if score >= 6:
            return "High"
        elif score >= 4:
            return "Medium"
        else:
            return "Low"
    
    def _generate_notes(self, app_type: str, has_aws: bool) -> str:
        """
        Generate migration notes based on app characteristics.
        """
        notes = []
        
        if "Windows Service" in app_type:
            notes.append("Requires Cloud Run or Compute Engine with agent")
        if "SOAP" in app_type:
            notes.append("Consider gRPC or REST API conversion")
        if "DLL" in app_type:
            notes.append("Analyze dependencies before migration")
        if has_aws:
            notes.append("AWS SDK replacement mapping needed")
        
        return "; ".join(notes) if notes else "Standard migration"
    
    def analyze_repos(self, repo_list: List[str], repo_owner: Optional[str] = None) -> List[Dict]:
        """
        Analyze a list of repositories, supporting both root and subfolder specs.
        """
        print(f"\n📊 Starting analysis of {len(repo_list)} repositories...\n")
        
        for repo_spec in repo_list:
            spec = self._parse_repo_spec(repo_spec)
            
            owner = spec["owner"] or repo_owner or self.org_or_user
            if not owner:
                print(f"  ⚠ Skipping {repo_spec}: no owner specified")
                continue
            
            result = self.analyze_repo(owner, spec["name"], branch=spec["branch"], subpath=spec["subpath"])
            if result:
                self.inventory.append(result)
        
        print(f"\n✅ Analysis complete. {len(self.inventory)} targets analyzed.\n")
        return self.inventory
    
    def export_to_excel(self, output_file: str = "migration_inventory.xlsx"):
        """
        Export inventory to Excel spreadsheet.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("⚠ openpyxl not installed. Run: pip install openpyxl")
            self.export_to_csv(output_file.replace(".xlsx", ".csv"))
            return
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Migration Inventory"
        
        # Headers
        headers = [
            "Repository Name",
            "Repository URL",
            "Description",
            "Application Type",
            "Target Framework",
            "Suggested GCP Service",
            "Database Engine",
            "Auth Architecture",
            "Logging Framework",
            "Secrets Management",
            "CI/CD Pipeline",
            "AWS SDK Detected",
            "AWS Services Used",
            "Hardcoded AWS Found",
            "Size (KB)",
            "Last Updated",
            "Is Archived",
            "Migration Complexity",
            "Migration Notes"
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data rows
        for item in self.inventory:
            ws.append([
                item.get("repo_name", "Unknown"),
                item.get("repo_url", ""),
                (item.get("description") or "")[:100],
                item.get("application_type", "Unknown"),
                item.get("target_framework", "Unknown"),
                item.get("suggested_gcp", "Compute Engine"),
                item.get("database_engine", "Unknown"),
                item.get("auth_strategy", "Default"),
                item.get("logging_framework", "Standard"),
                item.get("secrets_mgmt", "Environment"),
                item.get("cicd_pipeline", "None"),
                "Yes" if item.get("aws_sdk_detected") else "No",
                item.get("aws_services", "None"),
                "Yes" if item.get("has_hardcoded_aws") else "No",
                item.get("size_kb", 0),
                item.get("last_updated", ""),
                "Yes" if item.get("is_archived") else "No",
                item.get("migration_complexity", "Low"),
                item.get("notes", "")
            ])
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 12
        ws.column_dimensions["H"].width = 30
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 12
        ws.column_dimensions["K"].width = 18
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 18
        ws.column_dimensions["N"].width = 35
        
        # Add alternating row colors
        light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(self.inventory)+1), 1):
            if idx % 2 == 0:
                for cell in row:
                    cell.fill = light_fill
        
        try:
            wb.save(output_file)
            print(f"✅ Inventory exported to {output_file}")
        except PermissionError:
            print(f"❌ Error: Could not save '{output_file}'. Please make sure it's not open in Excel and try again.")
            alternate = f"migration_inventory_{int(datetime.now().timestamp())}.xlsx"
            try:
                wb.save(alternate)
                print(f"⚠ Saved to alternate file: {alternate}")
            except Exception:
                pass
        except Exception as e:
            print(f"❌ Error saving Excel: {e}")
    
    def export_to_csv(self, output_file: str = "migration_inventory.csv"):
        """
        Export inventory to CSV as fallback.
        """
        import csv
        
        fieldnames = [
            "repo_name", "repo_url", "description", "application_type",
            "target_framework", "suggested_gcp", "database_engine",
            "auth_strategy", "logging_framework", "secrets_mgmt",
            "cicd_pipeline", "aws_sdk_detected", "aws_services",
            "has_hardcoded_aws", "size_kb", "last_updated",
            "is_archived", "migration_complexity", "notes"
        ]
        
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.inventory)
        
        print(f"✅ Inventory exported to {output_file}")


# Example usage
if __name__ == "__main__":
    # Get GitHub token from environment or prompt
    GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")
    
    # List of repos to analyze (format: "owner/repo" or just "repo")
    REPOS_TO_ANALYZE = [
        "aws-samples/aws-net-guides/tree/master/SampleApplications/2022/MediaCatalog/MediaLibrary4.8/MediaLibrary",
    ]
    
    # Initialize analyzer
    analyzer = GitHubRepoAnalyzer(GITHUB_TOKEN, org_or_user="yourorg")
    
    # Analyze all repos
    results = analyzer.analyze_repos(REPOS_TO_ANALYZE)
    
    # Export to Excel
    analyzer.export_to_excel("AWS_to_GCP_Migration_Inventory.xlsx")
    
    # Optionally export summary
    print("\n📋 INVENTORY SUMMARY:")
    print(f"Total repos: {len(results)}")
    print(f"High complexity: {sum(1 for r in results if r['migration_complexity'] == 'High')}")
    print(f"Medium complexity: {sum(1 for r in results if r['migration_complexity'] == 'Medium')}")
    print(f"Low complexity: {sum(1 for r in results if r['migration_complexity'] == 'Low')}")
    print(f"AWS SDK detected: {sum(1 for r in results if r['aws_sdk_detected'])}")